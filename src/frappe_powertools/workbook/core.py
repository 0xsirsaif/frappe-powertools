"""Core workbook validation types and functions.

This module is Frappe-agnostic and contains pure Python implementations
for validating CSV/XLSX workbooks against Pydantic models.
"""

from __future__ import annotations

import csv
from collections.abc import Iterator
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, BinaryIO, Generic, Literal, Mapping, Optional, TextIO, Type, TypeVar

from pydantic import BaseModel, ConfigDict, ValidationError

# Type variable for Pydantic models
TModel = TypeVar("TModel", bound=BaseModel)


class TabularFormat(str, Enum):
    """Supported workbook formats."""
    
    auto = "auto"
    csv = "csv"
    xlsx = "xlsx"


@dataclass
class RowContext:
    """Metadata and raw data for a single workbook row."""
    
    row_index: int  # 1-based sheet row index (Excel/CSV row number)
    raw: Mapping[str, Any]  # header -> cell value (normalized)


@dataclass
class RowResult(Generic[TModel]):
    """Result of validating a single row.
    
    Attributes:
        context: Row metadata and raw data
        model: Validated Pydantic model instance (None if validation failed)
        error: Pydantic ValidationError (None if validation succeeded)
    """
    
    context: RowContext
    model: Optional[TModel]
    error: Optional[ValidationError]
    
    @property
    def is_valid(self) -> bool:
        """Check if the row validation succeeded."""
        return self.model is not None and self.error is None


@dataclass
class WorkbookSummary:
    """Summary statistics for workbook validation."""
    
    total_rows: int
    valid_rows: int
    invalid_rows: int
    
    @property
    def error_rate(self) -> float:
        """Calculate the error rate as a percentage."""
        if self.total_rows == 0:
            return 0.0
        return (self.invalid_rows / self.total_rows) * 100


@dataclass
class WorkbookValidationResult(Generic[TModel]):
    """Complete result of workbook validation.
    
    Attributes:
        summary: Validation statistics
        rows: List of individual row results
    """
    
    summary: WorkbookSummary
    rows: list[RowResult[TModel]]
    
    @property
    def valid_models(self) -> list[TModel]:
        """Get all successfully validated models."""
        return [row.model for row in self.rows if row.model is not None]
    
    @property
    def errors(self) -> list[tuple[int, ValidationError]]:
        """Get all validation errors with row indices."""
        return [(row.context.row_index, row.error) for row in self.rows if row.error is not None]


@dataclass
class WorkbookConfig:
    """Configuration for workbook parsing and validation."""
    
    format: TabularFormat = TabularFormat.auto
    header_row: int = 1  # 1-based row index for headers
    data_row_start: int | None = None  # default: header_row + 1 if None
    delimiter: str = ","  # CSV delimiter
    sheet_name: str | None = None  # default: first sheet for XLSX
    extra: Literal["ignore", "forbid", "allow"] = "ignore"
    stop_on_first_error: bool = False
    max_rows: int | None = None  # None means no explicit limit
    
    def __post_init__(self):
        """Set defaults for computed fields."""
        if self.data_row_start is None:
            self.data_row_start = self.header_row + 1
        
        # Validate configuration
        if self.header_row < 1:
            raise ValueError("header_row must be >= 1")
        if self.data_row_start < self.header_row:
            raise ValueError("data_row_start must be >= header_row")
        if self.max_rows is not None and self.max_rows < 1:
            raise ValueError("max_rows must be >= 1")


def _get_model_with_extra_config(
    model: Type[TModel],
    extra: Literal["ignore", "forbid", "allow"],
) -> Type[TModel]:
    """Get a model class with the specified extra field configuration.
    
    Args:
        model: Base Pydantic model class
        extra: How to handle extra fields ("ignore", "forbid", or "allow")
    
    Returns:
        Model class with the specified extra configuration
    """
    if extra == "forbid":
        class StrictModel(model):
            model_config = ConfigDict(extra="forbid")
        return StrictModel
    elif extra == "allow":
        class AllowModel(model):
            model_config = ConfigDict(extra="allow")
        return AllowModel
    else:
        # Use original model (which should have extra="ignore" by default)
        return model


def _normalize_value(value: Any) -> Any:
    """Normalize a cell value for Pydantic validation.
    
    - Strips whitespace from strings
    - Converts empty strings to None
    
    Args:
        value: Raw cell value
    
    Returns:
        Normalized value
    """
    if isinstance(value, str):
        clean_value = value.strip()
        # Convert empty strings to None for better Pydantic handling
        return None if clean_value == '' else clean_value
    return value


def _normalize_row_dict(row_dict: Mapping[str, Any]) -> dict[str, Any]:
    """Normalize a row dictionary for Pydantic validation.
    
    - Strips whitespace from keys and values
    - Converts empty strings to None
    
    Args:
        row_dict: Raw row dictionary
    
    Returns:
        Normalized row dictionary
    """
    normalized_row = {}
    for key, value in row_dict.items():
        # Strip whitespace from both keys and values
        clean_key = key.strip() if key else key
        normalized_row[clean_key] = _normalize_value(value)
    return normalized_row


def _detect_format(
    fp: BinaryIO | TextIO,
    file_name: str | None = None,
) -> TabularFormat:
    """Detect workbook format from file name or content.
    
    Args:
        fp: File-like object
        file_name: Optional filename to help with detection
        
    Returns:
        Detected TabularFormat (csv or xlsx)
    """
    # First, try to use file_name extension if available
    if file_name:
        file_name_lower = file_name.lower()
        if file_name_lower.endswith(('.xlsx', '.xlsm')):
            return TabularFormat.xlsx
        elif file_name_lower.endswith('.csv'):
            return TabularFormat.csv
        # For other extensions, fall through to content detection
    
    # Fallback to content-based detection
    if isinstance(fp, TextIO):
        # TextIO is always CSV
        return TabularFormat.csv
    
    # For binary streams, sniff the magic bytes
    # Check if it's a binary stream (not TextIO and has read method)
    is_binary = not isinstance(fp, TextIO) and hasattr(fp, 'read')
    if is_binary:
        # Save current position
        try:
            original_position = fp.tell()
        except (AttributeError, OSError):
            original_position = 0
        
        try:
            # Always read from the beginning for detection
            try:
                fp.seek(0)
            except (AttributeError, OSError):
                # If we can't seek, try reading from current position
                pass
            
            # Read first few bytes to check for ZIP magic (XLSX files are ZIP archives)
            sample = fp.read(4)
            
            # Always try to reset to original position
            try:
                fp.seek(original_position)
            except (AttributeError, OSError):
                # If we can't seek back, try to seek to 0 as fallback
                try:
                    fp.seek(0)
                except (AttributeError, OSError):
                    pass
            
            # XLSX files start with ZIP magic bytes: PK\x03\x04
            if len(sample) >= 4 and sample[:2] == b'PK' and sample[2:4] == b'\x03\x04':
                return TabularFormat.xlsx
            else:
                # Default to CSV for binary streams without ZIP magic
                return TabularFormat.csv
        except Exception:
            # If reading fails, try to reset and default to CSV
            try:
                fp.seek(original_position)
            except (AttributeError, OSError):
                try:
                    fp.seek(0)
                except (AttributeError, OSError):
                    pass
            return TabularFormat.csv
    
    # Default fallback
    return TabularFormat.csv


def iter_validated_rows(
    fp: BinaryIO | TextIO,
    model: Type[TModel],
    *,
    config: WorkbookConfig | None = None,
    file_name: str | None = None,
) -> Iterator[RowResult[TModel]]:
    """Stream rows from a CSV/XLSX file and validate each row with Pydantic.
    
    Args:
        fp: File-like object containing CSV or XLSX data
        model: Pydantic model class to validate each row against
        config: Configuration for parsing and validation
        file_name: Optional filename to help with format detection
    
    Yields:
        RowResult for each processed row
    
    Notes:
        - Uses config.format to choose CSV vs XLSX, or auto-detects when set to 'auto'
        - Yields RowResult for each row
        - Obeys stop_on_first_error and max_rows
    """
    if config is None:
        config = WorkbookConfig()
    
    # Determine format to use
    if config.format == TabularFormat.xlsx:
        format_to_use = TabularFormat.xlsx
    elif config.format == TabularFormat.csv:
        format_to_use = TabularFormat.csv
    elif config.format == TabularFormat.auto:
        # Auto-detect format from file_name or content
        format_to_use = _detect_format(fp, file_name)
        # Ensure stream is at position 0 after detection for parsing
        try:
            fp.seek(0)
        except (AttributeError, OSError):
            pass
    else:
        # Fallback to auto-detection for unknown formats
        format_to_use = _detect_format(fp, file_name)
        # Ensure stream is at position 0 after detection for parsing
        try:
            fp.seek(0)
        except (AttributeError, OSError):
            pass
    
    if format_to_use == TabularFormat.csv:
        yield from _iter_csv_rows(fp, model, config)
    elif format_to_use == TabularFormat.xlsx:
        yield from _iter_xlsx_rows(fp, model, config)


def _iter_csv_rows(
    fp: BinaryIO | TextIO,
    model: Type[TModel],
    config: WorkbookConfig,
) -> Iterator[RowResult[TModel]]:
    """Internal helper to parse and validate CSV rows."""
    import io
    
    # Ensure we have a text stream
    # Check if it's a binary stream by trying to read a small sample
    if hasattr(fp, 'mode') and 'b' in fp.mode:
        # It's a binary file
        content = fp.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        fp = io.StringIO(content)
    elif not hasattr(fp, 'mode'):
        # It might be BytesIO or similar - try to read and check
        original_position = fp.tell() if hasattr(fp, 'tell') else 0
        try:
            sample = fp.read(0)  # Try to read 0 bytes to check type
            fp.seek(original_position)
            if isinstance(sample, bytes):
                # It's binary
                content = fp.read()
                if isinstance(content, bytes):
                    content = content.decode('utf-8')
                fp = io.StringIO(content)
        except:
            # If anything fails, assume it's already text
            pass
    
    # Create CSV reader
    reader = csv.DictReader(fp, delimiter=config.delimiter)
    
    # Track row count (1-based, including header)
    current_row_index = 1  # Header is row 1
    rows_processed = 0
    
    # Configure Pydantic model validation based on config.extra
    model_to_use = _get_model_with_extra_config(model, config.extra)
    
    # Iterate through data rows
    for row_dict in reader:
        current_row_index += 1
        
        # Check if we're at the data start row yet
        if current_row_index < config.data_row_start:
            continue
        
        # Normalize row data: strip whitespace and convert empty strings to None
        normalized_row = _normalize_row_dict(row_dict)
        
        # Create row context with original data
        context = RowContext(row_index=current_row_index, raw=row_dict)
        
        # Try to validate the row with normalized data
        try:
            validated_model = model_to_use.model_validate(normalized_row)
            result = RowResult(context=context, model=validated_model, error=None)
        except ValidationError as e:
            result = RowResult(context=context, model=None, error=e)
        
        yield result
        
        rows_processed += 1
        
        # Check stopping conditions
        if config.stop_on_first_error and result.error is not None:
            break
        
        if config.max_rows is not None and rows_processed >= config.max_rows:
            break


def _iter_xlsx_rows(
    fp: BinaryIO | TextIO,
    model: Type[TModel],
    config: WorkbookConfig,
) -> Iterator[RowResult[TModel]]:
    """Internal helper to parse and validate XLSX rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX support. Install it with: pip install openpyxl"
        )
    
    # Ensure we have a binary stream for openpyxl
    if isinstance(fp, TextIO):
        raise ValueError("XLSX files require binary input. Please provide a BinaryIO stream.")
    
    # openpyxl needs the file to be seekable, so read all content if needed
    if not hasattr(fp, 'seek') or not hasattr(fp, 'tell'):
        # Read all content into memory
        content = fp.read()
        import io
        fp = io.BytesIO(content)
    
    # Load workbook in read-only mode for streaming
    wb = load_workbook(fp, read_only=True, data_only=True)
    
    # Select sheet
    if config.sheet_name:
        if config.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{config.sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")
        ws = wb[config.sheet_name]
    else:
        ws = wb.active
    
    # Read header row
    header_row_num = config.header_row
    header_cells = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))
    
    if not header_cells:
        wb.close()
        return
    
    headers = []
    for cell_value in header_cells[0]:
        if cell_value is not None:
            header_str = str(cell_value).strip()
            if header_str:  # Ignore empty headers
                headers.append(header_str)
    
    if not headers:
        wb.close()
        return
    
    # Configure Pydantic model validation based on config.extra
    model_to_use = _get_model_with_extra_config(model, config.extra)
    
    # Iterate through data rows
    data_start_row = config.data_row_start
    rows_processed = 0
    
    for row_num, row_values in enumerate(
        ws.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row
    ):
        # Check if row is completely empty (all None)
        if all(v is None for v in row_values):
            continue
        
        # Build row dictionary from headers and values
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row_values):
                value = row_values[i]
                # Normalize value (strip strings, convert empty strings to None)
                row_dict[header] = _normalize_value(value)
        
        # Create row context (1-based row index)
        context = RowContext(row_index=row_num, raw=row_dict)
        
        # Try to validate the row
        try:
            validated_model = model_to_use.model_validate(row_dict)
            result = RowResult(context=context, model=validated_model, error=None)
        except ValidationError as e:
            result = RowResult(context=context, model=None, error=e)
        
        yield result
        
        rows_processed += 1
        
        # Check stopping conditions
        if config.stop_on_first_error and result.error is not None:
            break
        
        if config.max_rows is not None and rows_processed >= config.max_rows:
            break
    
    wb.close()


def validate_workbook(
    fp: BinaryIO | TextIO,
    model: Type[TModel],
    *,
    config: WorkbookConfig | None = None,
    file_name: str | None = None,
) -> WorkbookValidationResult[TModel]:
    """Convenience wrapper around iter_validated_rows that collects all results.
    
    Args:
        fp: File-like object containing CSV or XLSX data
        model: Pydantic model class to validate each row against
        config: Configuration for parsing and validation
        file_name: Optional filename to help with format detection
    
    Returns:
        WorkbookValidationResult with summary and all row results
    """
    rows = list(iter_validated_rows(fp, model, config=config, file_name=file_name))
    
    # Calculate summary
    total_rows = len(rows)
    valid_rows = sum(1 for row in rows if row.is_valid)
    invalid_rows = total_rows - valid_rows
    
    summary = WorkbookSummary(
        total_rows=total_rows,
        valid_rows=valid_rows,
        invalid_rows=invalid_rows
    )
    
    return WorkbookValidationResult(summary=summary, rows=rows)
