"""Tests for XLSX workbook validation functionality."""

import io

import pytest
from openpyxl import Workbook
from pydantic import BaseModel, ConfigDict

from frappe_powertools.workbook.core import (
    WorkbookConfig,
    TabularFormat,
    iter_validated_rows,
    validate_workbook,
)


# Test models
class CustomerRow(BaseModel):
    """Sample model for testing."""

    code: str
    name: str
    email: str | None = None
    age: int | None = None


class StrictCustomerRow(BaseModel):
    """Model that forbids extra fields by default."""

    model_config = ConfigDict(extra="forbid")

    code: str
    name: str


# Fixtures
@pytest.fixture
def customer_model():
    """Fixture for CustomerRow model."""
    return CustomerRow


@pytest.fixture
def default_xlsx_config():
    """Fixture for default XLSX config."""
    return WorkbookConfig(format=TabularFormat.xlsx)


@pytest.fixture
def valid_xlsx_workbook():
    """Create a valid XLSX workbook in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Headers
    ws.append(["code", "name", "email", "age"])

    # Data rows
    ws.append(["CUST001", "Alice Smith", "alice@example.com", 30])
    ws.append(["CUST002", "Bob Jones", "bob@example.com", 25])
    ws.append(["CUST003", "Charlie Brown", None, 40])

    # Save to BytesIO
    fp = io.BytesIO()
    wb.save(fp)
    fp.seek(0)
    return fp


@pytest.fixture
def missing_field_xlsx_workbook():
    """Create XLSX workbook with missing required fields."""
    wb = Workbook()
    ws = wb.active

    ws.append(["code", "email", "age"])
    ws.append(["CUST001", "alice@example.com", 30])
    ws.append(["CUST002", "bob@example.com", 25])

    fp = io.BytesIO()
    wb.save(fp)
    fp.seek(0)
    return fp


@pytest.fixture
def extra_field_xlsx_workbook():
    """Create XLSX workbook with extra fields."""
    wb = Workbook()
    ws = wb.active

    ws.append(["code", "name", "email", "extra_field"])
    ws.append(["CUST001", "Alice Smith", "alice@example.com", "extra_value"])
    ws.append(["CUST002", "Bob Jones", "bob@example.com", "another_value"])

    fp = io.BytesIO()
    wb.save(fp)
    fp.seek(0)
    return fp


@pytest.fixture
def mixed_validity_xlsx_workbook():
    """Create XLSX workbook with mixed valid/invalid rows."""
    wb = Workbook()
    ws = wb.active

    ws.append(["code", "name", "age"])
    ws.append(["CUST001", "Alice", 30])
    ws.append(["CUST002", "Bob", "invalid_age"])
    ws.append(["CUST003", "Charlie", 25])
    ws.append(["CUST004", None, 35])
    ws.append(["CUST005", "Eve", 28])

    fp = io.BytesIO()
    wb.save(fp)
    fp.seek(0)
    return fp


@pytest.fixture
def empty_xlsx_workbook():
    """Create XLSX workbook with only headers."""
    wb = Workbook()
    ws = wb.active

    ws.append(["code", "name", "email"])

    fp = io.BytesIO()
    wb.save(fp)
    fp.seek(0)
    return fp


@pytest.fixture
def multi_sheet_xlsx_workbook():
    """Create XLSX workbook with multiple sheets."""
    wb = Workbook()

    # First sheet
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["code", "name"])
    ws1.append(["CUST001", "Alice"])

    # Second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["code", "name"])
    ws2.append(["CUST002", "Bob"])

    fp = io.BytesIO()
    wb.save(fp)
    fp.seek(0)
    return fp


# Helper functions
def assert_all_valid(results):
    """Assert all results are valid."""
    assert all(r.is_valid for r in results)
    assert all(r.model is not None for r in results)
    assert all(r.error is None for r in results)


def assert_all_invalid(results):
    """Assert all results are invalid."""
    assert all(not r.is_valid for r in results)
    assert all(r.model is None for r in results)
    assert all(r.error is not None for r in results)


# Tests
class TestIterValidatedRowsXLSX:
    """Test iter_validated_rows with XLSX input."""

    def test_all_valid(self, customer_model, valid_xlsx_workbook, default_xlsx_config):
        """Test with all valid XLSX data."""
        fp = valid_xlsx_workbook
        results = list(iter_validated_rows(fp, customer_model, config=default_xlsx_config))

        assert len(results) == 3
        assert_all_valid(results)

        # Check row indices (header is row 1, first data row is 2)
        assert results[0].context.row_index == 2
        assert results[1].context.row_index == 3
        assert results[2].context.row_index == 4

        # Check model data
        assert results[0].model.code == "CUST001"
        assert results[0].model.name == "Alice Smith"
        assert results[0].model.email == "alice@example.com"
        assert results[0].model.age == 30

        # Third row has None email
        assert results[2].model.email is None

    def test_missing_required_fields(
        self, customer_model, missing_field_xlsx_workbook, default_xlsx_config
    ):
        """Test with missing required fields."""
        fp = missing_field_xlsx_workbook
        results = list(iter_validated_rows(fp, customer_model, config=default_xlsx_config))

        assert len(results) == 2
        assert_all_invalid(results)

        # Check error mentions missing 'name' field
        for result in results:
            error_dict = result.error.errors()[0]
            assert error_dict["loc"] == ("name",)
            assert error_dict["type"] == "missing"

    @pytest.mark.parametrize(
        "extra,expected_valid",
        [
            ("forbid", False),
            ("ignore", True),
            ("allow", True),
        ],
    )
    def test_extra_columns(self, customer_model, extra_field_xlsx_workbook, extra, expected_valid):
        """Test extra columns with different extra settings."""
        fp = extra_field_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx, extra=extra)
        results = list(iter_validated_rows(fp, customer_model, config=config))

        assert len(results) == 2

        if expected_valid:
            assert_all_valid(results)
            if extra == "allow":
                # With allow, extra field should be in model
                assert hasattr(results[0].model, "extra_field")
            else:
                # With ignore, extra field should not be in model
                assert not hasattr(results[0].model, "extra_field")
        else:
            assert_all_invalid(results)
            # Check error mentions extra field
            error_dict = results[0].error.errors()[0]
            assert error_dict["type"] == "extra_forbidden"

    @pytest.mark.parametrize(
        "max_rows,expected_count",
        [
            (1, 1),
            (3, 3),
            (5, 5),
            (10, 5),  # More than available rows
        ],
    )
    def test_max_rows(self, customer_model, default_xlsx_config, max_rows, expected_count):
        """Test max_rows configuration."""
        # Create workbook with 5 rows for this test
        wb = Workbook()
        ws = wb.active
        ws.append(["code", "name"])
        ws.append(["CUST001", "Alice"])
        ws.append(["CUST002", "Bob"])
        ws.append(["CUST003", "Charlie"])
        ws.append(["CUST004", "David"])
        ws.append(["CUST005", "Eve"])

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        config = WorkbookConfig(format=TabularFormat.xlsx, max_rows=max_rows)
        results = list(iter_validated_rows(fp, customer_model, config=config))

        assert len(results) == expected_count
        if results:
            assert results[0].model.code == "CUST001"

    def test_stop_on_first_error(
        self, customer_model, mixed_validity_xlsx_workbook, default_xlsx_config
    ):
        """Test stop_on_first_error configuration."""
        fp = mixed_validity_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx, stop_on_first_error=True)
        results = list(iter_validated_rows(fp, customer_model, config=config))

        # Should stop after second row (first error)
        assert len(results) == 2
        assert results[0].is_valid
        assert not results[1].is_valid

    def test_empty_xlsx(self, customer_model, empty_xlsx_workbook, default_xlsx_config):
        """Test with empty XLSX (only headers)."""
        fp = empty_xlsx_workbook
        results = list(iter_validated_rows(fp, customer_model, config=default_xlsx_config))

        assert len(results) == 0

    def test_sheet_selection(self, customer_model, multi_sheet_xlsx_workbook, default_xlsx_config):
        """Test selecting specific sheet by name."""
        fp = multi_sheet_xlsx_workbook

        # Test Sheet1
        config1 = WorkbookConfig(format=TabularFormat.xlsx, sheet_name="Sheet1")
        results1 = list(iter_validated_rows(fp, customer_model, config=config1))
        assert len(results1) == 1
        assert results1[0].model.code == "CUST001"
        assert results1[0].model.name == "Alice"

        # Reset file pointer
        fp.seek(0)

        # Test Sheet2
        config2 = WorkbookConfig(format=TabularFormat.xlsx, sheet_name="Sheet2")
        results2 = list(iter_validated_rows(fp, customer_model, config=config2))
        assert len(results2) == 1
        assert results2[0].model.code == "CUST002"
        assert results2[0].model.name == "Bob"

    def test_sheet_not_found(self, customer_model, valid_xlsx_workbook, default_xlsx_config):
        """Test error when sheet name doesn't exist."""
        fp = valid_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx, sheet_name="NonExistentSheet")

        with pytest.raises(ValueError, match="Sheet 'NonExistentSheet' not found"):
            list(iter_validated_rows(fp, customer_model, config=config))

    def test_custom_header_and_data_rows(self, customer_model, default_xlsx_config):
        """Test custom header_row and data_row_start."""
        wb = Workbook()
        ws = wb.active

        # Add some comment rows
        ws.append(["Comment line 1"])
        ws.append(["Comment line 2"])

        # Header on row 3
        ws.append(["code", "name", "email"])

        # Another comment
        ws.append(["Another comment"])

        # Data starts on row 5
        ws.append(["CUST001", "Alice Smith", "alice@example.com"])
        ws.append(["CUST002", "Bob Jones", "bob@example.com"])

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        config = WorkbookConfig(format=TabularFormat.xlsx, header_row=3, data_row_start=5)

        results = list(iter_validated_rows(fp, customer_model, config=config))

        assert len(results) == 2
        assert_all_valid(results)
        assert results[0].context.row_index == 5
        assert results[1].context.row_index == 6

    def test_empty_rows_skipped(self, customer_model, default_xlsx_config):
        """Test that completely empty rows are skipped."""
        wb = Workbook()
        ws = wb.active

        ws.append(["code", "name"])
        ws.append(["CUST001", "Alice"])
        ws.append([None, None])  # Empty row
        ws.append(["CUST002", "Bob"])
        ws.append([None, None])  # Another empty row
        ws.append(["CUST003", "Charlie"])

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        results = list(iter_validated_rows(fp, customer_model, config=default_xlsx_config))

        # Should only have 3 rows (empty rows skipped)
        assert len(results) == 3
        assert_all_valid(results)
        assert results[0].model.code == "CUST001"
        assert results[1].model.code == "CUST002"
        assert results[2].model.code == "CUST003"


class TestValidateWorkbookXLSX:
    """Test validate_workbook with XLSX input."""

    def test_summary_statistics(self, customer_model, mixed_validity_xlsx_workbook):
        """Test summary with mixed valid/invalid rows."""
        fp = mixed_validity_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx)
        result = validate_workbook(fp, customer_model, config=config)

        assert result.summary.total_rows == 5
        assert result.summary.valid_rows == 3  # Rows 1, 3, 5
        assert result.summary.invalid_rows == 2  # Row 2 (invalid age), Row 4 (empty name)
        assert result.summary.error_rate == 40.0

        # Check valid models
        valid_models = result.valid_models
        assert len(valid_models) == 3
        assert valid_models[0].code == "CUST001"
        assert valid_models[1].code == "CUST003"
        assert valid_models[2].code == "CUST005"

        # Check errors
        errors = result.errors
        assert len(errors) == 2
        assert errors[0][0] == 3  # Row 3 (Bob with invalid age)
        assert errors[1][0] == 5  # Row 5 (missing name)

    def test_all_valid(self, customer_model, valid_xlsx_workbook):
        """Test with all valid rows."""
        fp = valid_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx)
        result = validate_workbook(fp, customer_model, config=config)

        assert result.summary.total_rows == 3
        assert result.summary.valid_rows == 3
        assert result.summary.invalid_rows == 0
        assert result.summary.error_rate == 0.0

    def test_all_invalid(self, customer_model, missing_field_xlsx_workbook):
        """Test with all invalid rows."""
        fp = missing_field_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx)
        result = validate_workbook(fp, customer_model, config=config)

        assert result.summary.total_rows == 2
        assert result.summary.valid_rows == 0
        assert result.summary.invalid_rows == 2
        assert result.summary.error_rate == 100.0

    def test_empty_workbook(self, customer_model, empty_xlsx_workbook):
        """Test with empty workbook."""
        fp = empty_xlsx_workbook
        config = WorkbookConfig(format=TabularFormat.xlsx)
        result = validate_workbook(fp, customer_model, config=config)

        assert result.summary.total_rows == 0
        assert result.summary.valid_rows == 0
        assert result.summary.invalid_rows == 0
        assert result.summary.error_rate == 0.0
        assert len(result.valid_models) == 0
        assert len(result.errors) == 0
